from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory
import os
import random
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key'

excel_file = 'annotations.xlsx'
current_image = None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start', methods=['POST'])
def start():
    directory = request.form['directory']
    categories_input = request.form['categories']

    # Validate directory
    if not os.path.isdir(directory):
        return "Invalid directory. Please go back and enter a valid directory."

    # Get list of image files in the directory
    allowed_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
    images_list = [f for f in os.listdir(directory)
                   if os.path.splitext(f)[1].lower() in allowed_extensions]

    if not images_list:
        return "No images found in the directory. Please go back and enter a directory with images."

    # Store the directory and images_list in session
    session['directory'] = directory
    session['images_list'] = images_list

    # Parse categories
    categories = [c.strip() for c in categories_input.split(',')]
    session['categories'] = categories

    # Initialize Excel file if it doesn't exist
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Annotations'
        # Write headers
        headers = ['Image'] + categories
        ws.append(headers)
        wb.save(excel_file)
    else:
        # Load existing annotations to skip already annotated images
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Annotations']
        annotated_images = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            annotated_images.add(row[0])
        session['images_list'] = [img for img in images_list if img not in annotated_images]

    return redirect(url_for('annotate'))

@app.route('/annotate', methods=['GET', 'POST'])
def annotate():
    global current_image

    directory = session.get('directory')
    categories = session.get('categories')
    images_list = session.get('images_list')

    if not images_list:
        return redirect(url_for('finished'))

    if request.method == 'POST':
        # Save annotations
        selected_categories = request.form.getlist('categories')
        image_name = current_image

        # Save to Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Annotations']
        row = [image_name]
        for cat in categories:
            row.append('1' if cat in selected_categories else '0')
        ws.append(row)
        wb.save(excel_file)

        # Remove the image from images_list
        images_list.remove(current_image)
        session['images_list'] = images_list

        # Check if there are images left
        if not images_list:
            return redirect(url_for('finished'))

    # Get next image
    if images_list:
        current_image = random.choice(images_list)
        image_url = url_for('image', filename=current_image)
        return render_template('annotate.html', image_url=image_url, categories=categories)
    else:
        return redirect(url_for('finished'))

@app.route('/image/<filename>')
def image(filename):
    directory = session.get('directory')
    return send_from_directory(directory, filename)

@app.route('/finished')
def finished():
    return render_template('finished.html')

if __name__ == '__main__':
    app.run(debug=True)
